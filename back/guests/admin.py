from django.utils.text import slugify
from openpyxl import load_workbook
from django import forms
from django.contrib import admin, messages
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse

from .models import Guest


class GuestImportForm(forms.Form):
    file = forms.FileField(label='Arquivo Excel (.xlsx)')


@admin.register(Guest)
class GuestAdmin(admin.ModelAdmin):
    list_display = ('name', 'family_head', 'phone', 'response')
    list_filter = ('response', 'family_head')
    search_fields = ('name', 'phone', 'notes')
    change_list_template = 'admin/guests/guest/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-xlsx/', self.admin_site.admin_view(self.import_xlsx_view), name='guests_guest_import_xlsx'),
        ]
        return custom_urls + urls

    def import_xlsx_view(self, request):
        if request.method == 'POST':
            form = GuestImportForm(request.POST, request.FILES)
            if form.is_valid():
                uploaded_file = request.FILES['file']
                workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
                sheet = workbook.active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    messages.error(request, 'O arquivo enviado está vazio.')
                    return HttpResponseRedirect(reverse('admin:guests_guest_changelist'))

                columns = [str(value).strip() if value is not None else '' for value in rows[0]]
                if not columns:
                    messages.error(request, 'Não foi possível identificar as colunas do arquivo.')
                    return HttpResponseRedirect(reverse('admin:guests_guest_changelist'))

                def header_index(name):
                    for idx, column in enumerate(columns):
                        if column.lower() == name.lower():
                            return idx
                    return -1

                index_id = header_index('ID')
                index_name = header_index('Convidados')
                index_chefe = header_index('Chefe')
                index_telefone = header_index('Telefone')
                index_obs = header_index('Observação')

                if index_name == -1:
                    messages.error(request, 'A coluna "Convidados" é obrigatória no arquivo XLSX.')
                    return HttpResponseRedirect(reverse('admin:guests_guest_changelist'))

                def normalize_name(value):
                    if value is None:
                        return ''
                    return ' '.join(str(value).strip().split())

                def row_value(row, index):
                    if index < 0 or index >= len(row):
                        return ''
                    value = row[index]
                    return '' if value is None else normalize_name(value)

                def find_or_create_guest(name, source_id=''):
                    normalized = normalize_name(name)
                    if source_id:
                        source_key = slugify(str(source_id))
                        guest = Guest.objects.filter(name__iexact=normalized).filter(slug__icontains=source_key).first()
                        if guest is not None:
                            return guest

                    guest = Guest.objects.filter(name__iexact=normalized).order_by('id').first()
                    if guest is not None:
                        return guest

                    return Guest.objects.create(name=normalized)

                def fix_duplicate_head_slugs():
                    for guest in Guest.objects.filter(family_head__isnull=True).order_by('id'):
                        if not guest.slug:
                            guest.slug = guest.generate_unique_slug()
                            guest.save(update_fields=['slug'])

                fix_duplicate_head_slugs()

                created_map = {}
                created_by_id = {}
                created_by_name = {}

                for row in rows[1:]:
                    if not row:
                        continue

                    name = row_value(row, index_name)
                    if not name:
                        continue

                    source_id = row_value(row, index_id)
                    key = (source_id or name).lower()
                    if key not in created_map:
                        created_map[key] = find_or_create_guest(name, source_id)
                    if source_id:
                        created_by_id[source_id.lower()] = created_map[key]
                    created_by_name[name.lower()] = created_map[key]

                    guest = created_map[key]
                    guest.phone = row_value(row, index_telefone)
                    guest.notes = row_value(row, index_obs)
                    guest.family_head = None
                    if not guest.slug:
                        guest.slug = guest.generate_unique_slug(source_id=source_id)
                    guest.save()

                with transaction.atomic():
                    for row in rows[1:]:
                        if not row:
                            continue

                        name = row_value(row, index_name)
                        if not name:
                            continue

                        source_id = row_value(row, index_id)
                        chefe = row_value(row, index_chefe)
                        guest_key = (source_id or name).lower()
                        guest = created_map.get(guest_key)
                        if guest is None:
                            continue

                        family_head = None
                        if chefe:
                            chefe_ref = chefe.lower()
                            family_head = created_by_id.get(chefe_ref)
                            if family_head is None:
                                family_head = created_by_name.get(chefe_ref)
                            if family_head is None:
                                family_head = find_or_create_guest(chefe)
                                created_by_name[chefe_ref] = family_head

                        if family_head is not None:
                            guest.family_head = family_head
                        else:
                            guest.family_head = None

                        if source_id and guest.slug in (None, ''):
                            guest.slug = guest.generate_unique_slug(source_id=source_id)
                        elif not guest.slug:
                            guest.slug = guest.generate_unique_slug()
                        guest.save()

                messages.success(request, f'Arquivo importado com sucesso. {len(created_map)} convidados foram processados.')
                return HttpResponseRedirect(reverse('admin:guests_guest_changelist'))

        form = GuestImportForm()
        context = {
            'opts': self.model._meta,
            'form': form,
            'title': 'Importar convidados via XLSX',
        }
        return render(request, 'admin/guests/guest/import_xlsx.html', context)